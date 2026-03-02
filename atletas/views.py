from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db.models import Sum, Count
from django.http import FileResponse, HttpResponse
from django.conf import settings
from django.http import JsonResponse

from datetime import date, datetime
import os
import io
import calendar
import pandas as pd

from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import Atleta, Mensualidad, Equipo, Administrador
from .forms import AtletaForm, EquipoForm, AdministradorForm

def solo_login(view_func):
    return login_required(view_func)

@solo_login
def lista_atletas(request):
    atletas = Atleta.objects.all()

    categoria = request.GET.get('categoria')
    sexo = request.GET.get('sexo')
    posicion = request.GET.get('posicion')
    turno = request.GET.get('turno')
    numero = request.GET.get('numero')
    cedula = request.GET.get('cedula')

    if categoria:
        atletas = atletas.filter(categoria=categoria)
    if sexo:
        atletas = atletas.filter(sexo=sexo)
    if posicion:
        atletas = atletas.filter(posicion=posicion)
    if turno:
        atletas = atletas.filter(turno=turno)
    if numero:
        atletas = atletas.filter(numero_camisa=numero)
    if cedula:
        atletas = atletas.filter(cedula__icontains=cedula)

    atletas = atletas.order_by("apellido", "nombre", "id")
    
    # paginar
    paginator = Paginator(atletas, 8)
    page_number = request.GET.get('page')
    atletas = paginator.get_page(page_number)

    # para los selects
    categorias = Atleta.objects.values_list('categoria', flat=True).distinct()
    posiciones = Atleta.POSICIONES
    turnos = Atleta.TURNOS
    sexos = Atleta.SEXO


    return render(request, 'atletas/lista_atletas.html', {
        'atletas': atletas,
        'categorias': categorias,
        'posiciones': posiciones,
        'turnos': turnos,
        'sexos': sexos,
        'valores': request.GET,
    })

@solo_login
def agregar_atleta(request):
    if request.method == 'POST':
        form = AtletaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_atletas')
    else:
        form = AtletaForm()

    return render(request, 'atletas/agregar_atleta.html', {'form': form})

@solo_login
def detalle_atleta(request, atleta_id):
    atleta = get_object_or_404(Atleta, id=atleta_id)

    # ✅ Calcular despegue si aplica
    despegue = None
    if atleta.salto is not None and atleta.alcance is not None:
        despegue = round((atleta.salto - atleta.alcance) * 100, 2)

    context = {
        'atleta': atleta,
        'despegue': despegue,
    }
    return render(request, 'atletas/detalle_atleta.html', context)

@solo_login
def editar_atleta(request, atleta_id):
    atleta = get_object_or_404(Atleta, id=atleta_id)

    if request.method == 'POST':
        form = AtletaForm(request.POST, instance=atleta)
        if form.is_valid():
            form.save()
            return redirect('detalle_atleta', atleta_id=atleta.id)
    else:
        form = AtletaForm(instance=atleta)

    return render(request, 'atletas/editar_atleta.html', {
        'form': form,
        'atleta': atleta,
    })


@solo_login
def eliminar_atleta(request, atleta_id):
    atleta = get_object_or_404(Atleta, id=atleta_id)
    if request.method == 'POST':
        atleta.delete()
        return redirect('lista_atletas')
    return render(request, 'atletas/eliminar_atleta.html', {'atleta': atleta})


@solo_login
@require_POST
def actualizar_mensualidad(request, mensualidad_id):
    mensualidad = get_object_or_404(Mensualidad, id=mensualidad_id)

    def parse_bool(v):
        return str(v).strip().lower() in ("1", "true", "on", "yes")

    monto_raw = (request.POST.get("monto") or "").strip()
    try:
        monto = float(monto_raw) if monto_raw else 0.0
    except ValueError:
        monto = 0.0

    # ✅ aquí el fix
    exonerado = parse_bool(request.POST.get("exonerado", "false"))

    mensualidad.monto_pagado = monto
    mensualidad.exonerado = exonerado
    mensualidad.save()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        año = int(request.POST.get("año", date.today().year))
        mes = int(request.POST.get("mes", date.today().month))
        categoria = (request.POST.get("categoria") or "").strip()

        mensualidades_mes = Mensualidad.objects.filter(año=año, mes=mes)
        if categoria:
            mensualidades_mes = mensualidades_mes.filter(atleta__categoria=categoria)

        total_recaudado = float(mensualidades_mes.aggregate(total=Sum("monto_pagado"))["total"] or 0)
        total_atletas = mensualidades_mes.count()

        al_dia = 0
        exonerados = 0
        for m in mensualidades_mes:
            if m.exonerado:
                exonerados += 1
            elif float(m.monto_pagado) >= 8.0:
                al_dia += 1

        deudores = total_atletas - al_dia - exonerados

        if mensualidad.exonerado:
            estado = "exonerado"
        elif float(mensualidad.monto_pagado) >= 8.0:
            estado = "pagado"
        elif float(mensualidad.monto_pagado) > 0:
            estado = "parcial"
        else:
            estado = "no_pagado"

        return JsonResponse({
            "ok": True,
            "estado": estado,
            "monto": float(mensualidad.monto_pagado),
            "exonerado": bool(mensualidad.exonerado),
            "total_recaudado": total_recaudado,
            "al_dia": al_dia,
            "exonerados": exonerados,
            "deudores": deudores,
            "total_atletas": total_atletas,
        })

    return redirect("administracion")

@solo_login
def administracion(request):
    mes_actual = date.today().month
    año_actual = int(request.GET.get("año", date.today().year))
    mes_filtro = int(request.GET.get("mes", mes_actual))
    categoria_filtro = request.GET.get("categoria", "").strip()

    atletas = Atleta.objects.all()
    if categoria_filtro:
        atletas = atletas.filter(categoria=categoria_filtro)

    atletas = atletas.order_by("apellido", "nombre", "id")

    # =========================
    # 1) Asegurar mensualidades del año seleccionado (solo faltantes) con BULK
    # =========================
    atleta_ids = list(atletas.values_list("id", flat=True))

    existentes = set(
        Mensualidad.objects.filter(atleta_id__in=atleta_ids, año=año_actual)
        .values_list("atleta_id", "mes")
    )

    crear = []
    for atleta_id in atleta_ids:
        for mes in range(1, 13):
            if (atleta_id, mes) not in existentes:
                crear.append(
                    Mensualidad(
                        atleta_id=atleta_id,
                        año=año_actual,
                        mes=mes,
                        monto_pagado=0,
                        exonerado=False,
                    )
                )

    if crear:
        Mensualidad.objects.bulk_create(crear, ignore_conflicts=True)

    # =========================
    # 2) Preparar tabla
    # =========================
    mensualidades_dict = {
        m.atleta_id: m
        for m in Mensualidad.objects.filter(atleta_id__in=atleta_ids, año=año_actual, mes=mes_filtro)
    }

    registros = []
    for atleta in atletas:
        registros.append({
            "atleta": atleta,
            "mensualidad": mensualidades_dict.get(atleta.id)  # siempre debería existir ya
        })

    # =========================
    # 3) Resumen mensual
    # =========================
    mensualidades_mes = Mensualidad.objects.filter(mes=mes_filtro, año=año_actual)
    if categoria_filtro:
        mensualidades_mes = mensualidades_mes.filter(atleta__categoria=categoria_filtro)

    total_recaudado = mensualidades_mes.aggregate(total=Sum("monto_pagado"))["total"] or 0
    total_atletas = mensualidades_mes.count()

    al_dia = 0
    exonerados = 0
    for m in mensualidades_mes:
        if m.exonerado:
            exonerados += 1
        elif float(m.monto_pagado) >= 8.0:
            al_dia += 1

    deudores = total_atletas - al_dia - exonerados

    # =========================
    # 4) Paginación
    # =========================
    paginator = Paginator(registros, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    meses = [(i, date(1900, i, 1).strftime('%B')) for i in range(1, 13)]
    categorias = Atleta.objects.values_list("categoria", flat=True).distinct()

    context = {
        "page_obj": page_obj,
        "mes_filtro": mes_filtro,
        "mes_actual": mes_actual,
        "año_actual": año_actual,
        "categorias": categorias,
        "categoria_filtro": categoria_filtro,
        "meses": meses,
        "total_recaudado": total_recaudado,
        "al_dia": al_dia,
        "exonerados": exonerados,
        "deudores": deudores,
        "total_atletas": total_atletas,
    }
    return render(request, "atletas/administracion.html", context)

def calcular_edad(nacimiento):
    hoy = date.today()
    return hoy.year - nacimiento.year - ((hoy.month, hoy.day) < (nacimiento.month, nacimiento.day))

@solo_login
def registrar_equipo(request):
    edad_tope = request.GET.get('edad_tope') or request.POST.get('edad_tope')
    sexo = request.GET.get('sexo') or request.POST.get('sexo')
    atletas_filtrados = []
    mostrar_formulario = False
    error_maximo = False

    if edad_tope and sexo:
        try:
            edad_tope = int(edad_tope)
            todos = Atleta.objects.all()
            for atleta in todos:
                edad = calcular_edad(atleta.fecha_nacimiento)
                if edad <= edad_tope:
                    if sexo == 'mixto' or atleta.sexo == sexo:
                        atleta.edad = edad  # atributo temporal
                        atletas_filtrados.append(atleta)
            mostrar_formulario = True
        except ValueError:
            edad_tope = None

    if request.method == 'POST':
        form = EquipoForm(request.POST)
        atletas_ids = request.POST.getlist('atletas_seleccionados')

        # Validamos que no pasen de 14 atletas
        if len(atletas_ids) > 14:
            error_maximo = True
            mostrar_formulario = True  # Mantener el formulario visible
        elif form.is_valid():
            equipo = form.save(commit=False)
            equipo.sexo_equipo = sexo
            equipo.edad_tope = int(edad_tope) if edad_tope else None
            equipo.save()
            equipo.atletas.set(atletas_ids)
            return redirect('listar_equipos')

    else:
        form = EquipoForm()

    context = {
        'form': form,
        'edad_tope': edad_tope,
        'sexo': sexo,
        'atletas_disponibles': atletas_filtrados,
        'atletas_seleccionados': request.POST.getlist('atletas_seleccionados') if request.method == 'POST' else [],
        'mostrar_formulario': mostrar_formulario,
        'error_maximo': error_maximo
    }
    return render(request, 'equipos/registrar_equipo.html', context)


# Vista para listar equipos
@solo_login
def listar_equipos(request):
    equipos = Equipo.objects.all().order_by('nombre')

    nombre = request.GET.get('nombre', '').strip()

    if nombre:
        equipos = equipos.filter(nombre__icontains=nombre)

    paginator = Paginator(equipos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'valores': {
            'nombre': nombre,
        }
    }
    return render(request, 'equipos/listar_equipos.html', context)

@solo_login
def detalle_equipo(request, equipo_id):
    equipo = get_object_or_404(Equipo, pk=equipo_id)
    atletas = equipo.atletas.all().order_by('nombre')

    # Calcular edad para cada atleta
    for atleta in atletas:
        atleta.edad = calcular_edad(atleta.fecha_nacimiento)

    return render(request, 'equipos/detalle_equipo.html', {
        'equipo': equipo,
        'atletas': atletas
    })

@solo_login
def eliminar_equipo(request, equipo_id):
    equipo = get_object_or_404(Equipo, pk=equipo_id)

    if request.method == 'POST':
        nombre = equipo.nombre
        equipo.delete()
        return redirect('listar_equipos')

    return render(request, 'equipos/eliminar_equipo.html', {'equipo': equipo})


@solo_login
def editar_equipo(request, equipo_id):
    equipo = get_object_or_404(Equipo, pk=equipo_id)
    error_msg = None
    atletas_filtrados = []
    atletas_ids = []

    # Obtener sexo del equipo original (campo fijo)
    # Al inicio
    sexo = request.POST.get('sexo_equipo') or request.GET.get('sexo') or equipo.sexo_equipo


    # Obtener edad tope desde GET, POST o calcular según atletas actuales
    # Obtener edad_tope como string
    edad_tope = request.GET.get('edad_tope') or request.POST.get('edad_tope')

    # Si no hay edad_tope enviada, usar la que se guardó en el modelo
    if not edad_tope:
        if equipo.edad_tope:
            edad_tope = equipo.edad_tope
        elif equipo.atletas.exists():
            edad_tope = max([calcular_edad(a.fecha_nacimiento) for a in equipo.atletas.all()])
        else:
            edad_tope = ''

    # Validar edad_tope
    error_msg = None
    try:
        edad_tope_int = int(edad_tope)
        if edad_tope_int <= 0:
            raise ValueError
    except ValueError:
        edad_tope_int = None
        error_msg = "Edad tope inválida."

    # Si la edad_tope es válida, filtrar atletas
    if edad_tope_int:
        atletas_filtrados = []
        for atleta in Atleta.objects.all():
            edad = calcular_edad(atleta.fecha_nacimiento)
            if edad <= edad_tope_int and (sexo == 'mixto' or atleta.sexo == sexo):
                atleta.edad = edad
                atletas_filtrados.append(atleta)


    if request.method == 'POST':
        form = EquipoForm(request.POST, instance=equipo)
        atletas_ids = request.POST.getlist('atletas_seleccionados')

        if len(atletas_ids) > 14:
            error_msg = "Solo se pueden seleccionar hasta 14 atletas."
        elif form.is_valid():
            equipo = form.save(commit=False)
            equipo.sexo_equipo = sexo  # Se mantiene sin editar
            equipo.edad_tope = edad_tope_int
            equipo.save()
            equipo.atletas.set(atletas_ids)
            return redirect('detalle_equipo', equipo.id)
    else:
        form = EquipoForm(instance=equipo)
        atletas_ids = list(equipo.atletas.values_list('id', flat=True))

    context = {
        'form': form,
        'equipo': equipo,
        'sexo': sexo,
        'edad_tope': edad_tope,
        'atletas_disponibles': atletas_filtrados,
        'atletas_seleccionados': atletas_ids,
        'error_msg': error_msg,
        'mostrar_formulario': True  # Siempre mostrar el formulario
    }
    return render(request, 'equipos/editar_equipo.html', context)

@solo_login
def reporte_pagos_view(request):
    año_actual = datetime.now().year
    año = int(request.GET.get('año', año_actual))
    categoria = request.GET.get('categoria')
    cedula = request.GET.get('cedula', '').strip()
    nombre = request.GET.get('nombre', '').strip()
    mes_inicio = int(request.GET.get('mes_inicio') or 1)
    mes_fin = int(request.GET.get('mes_fin') or 12)
    estado_filtro = request.GET.get('estado', '').strip()  # ✅ limpio

    atletas = Atleta.objects.all()

    if categoria:
        atletas = atletas.filter(categoria=categoria)
    if nombre:
        atletas = atletas.filter(nombre__icontains=nombre)

    atletas_filtrados_por_cedula = False
    if cedula:
        atletas = atletas.filter(cedula__icontains=cedula)
        if atletas.count() == 1:
            atletas_filtrados_por_cedula = True

    resumen = []
    for atleta in atletas:
        pagos = []
        total_pagado = 0
        total_pendiente = 0
        tiene_estado = False

        for mes in range(mes_inicio, mes_fin + 1):
            mensualidad = Mensualidad.objects.filter(atleta=atleta, año=año, mes=mes).first()

            if mensualidad:
                monto = float(mensualidad.monto_pagado)
                if mensualidad.exonerado:
                    estado_mes = "exonerado"
                elif monto >= 8.00:
                    estado_mes = "pagado"
                elif monto > 0:
                    estado_mes = "parcial"
                else:
                    estado_mes = "no_pagado"
            else:
                monto = 0.00
                estado_mes = "no_pagado"

            pagos.append({'mes': mes, 'monto': monto, 'estado': estado_mes})

            # Totales (no se suma pendiente si es exonerado)
            if estado_mes != "exonerado":
                total_pagado += monto
                if monto < 8:
                    total_pendiente += (8 - monto)

            # ✅ filtro por estado (ahora claro y sin pisar variables)
            if estado_filtro and estado_mes == estado_filtro:
                tiene_estado = True

        # Si se está filtrando por estado, y no tiene ese estado, lo saco
        if estado_filtro and not tiene_estado and not atletas_filtrados_por_cedula:
            continue

        resumen.append({
            'id': atleta.id,
            'nombre': atleta.nombre,
            'apellido': atleta.apellido,
            'cedula': atleta.cedula,
            'categoria': atleta.categoria,
            'pagos': pagos,
            'total_pagado': total_pagado,
            'total_pendiente': total_pendiente
        })

    paginator = Paginator(resumen, 15)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    meses = [(i, calendar.month_name[i]) for i in range(mes_inicio, mes_fin + 1)]
    categorias = Atleta.objects.values_list('categoria', flat=True).distinct()

    context = {
        'resumen': page_obj,
        'categorias': categorias,
        'año_actual': año_actual,
        'año_seleccionado': año,
        'categoria_seleccionada': categoria,
        'cedula': cedula,
        'nombre': nombre,
        'mes_inicio': mes_inicio,
        'mes_fin': mes_fin,
        'estado': estado_filtro,
        'meses': meses,
        'page_obj': page_obj,
    }

    return render(request, 'reportes/reporte_pagos.html', context)

@solo_login
def exportar_pagos_excel(request):

    año = int(request.GET.get('año', date.today().year))
    categoria = request.GET.get('categoria')
    nombre = request.GET.get('nombre')
    cedula = request.GET.get('cedula')
    estado = request.GET.get('estado')  # pagado, pendiente, parcial, exonerado
    mes_inicio = int(request.GET.get('mes_inicio', 1))
    mes_fin = int(request.GET.get('mes_fin', 12))

    atletas = Atleta.objects.all()
    if categoria:
        atletas = atletas.filter(categoria=categoria)
    if nombre:
        atletas = atletas.filter(nombre__icontains=nombre)

    atletas_filtrados_por_cedula = False
    if cedula:
        atletas = atletas.filter(cedula__icontains=cedula)
        if atletas.count() == 1:
            atletas_filtrados_por_cedula = True


    data = []
    meses = [(i, date(1900, i, 1).strftime('%b')) for i in range(mes_inicio, mes_fin + 1)]

    for atleta in atletas:
        fila = {
            'Nombre': atleta.nombre,
            'Cédula': atleta.cedula,
            'Categoría': atleta.categoria,
        }
        total_pagado = 0
        total_pendiente = 0
        tiene_parcial = False
        tiene_exonerado = False
        todos_pagados = True

        for mes_num, mes_nombre in meses:
            mensualidad = Mensualidad.objects.filter(atleta=atleta, año=año, mes=mes_num).first()
            if mensualidad:
                monto = float(mensualidad.monto_pagado)
                if mensualidad.exonerado:
                    fila[mes_nombre] = "E"
                    tiene_exonerado = True
                elif monto >= 8:
                    fila[mes_nombre] = monto
                elif monto > 0:
                    fila[mes_nombre] = monto
                    tiene_parcial = True
                    todos_pagados = False
                else:
                    fila[mes_nombre] = 0
                    todos_pagados = False
                total_pagado += monto if not mensualidad.exonerado else 0
                total_pendiente += 0 if (monto >= 8 or mensualidad.exonerado) else (8 - monto)
            else:
                fila[mes_nombre] = 0
                total_pendiente += 8
                todos_pagados = False

        # Nueva lógica de filtro por estado (igual al PDF)
        tiene_estado = False

        for mes_num, _ in meses:
            mensualidad = Mensualidad.objects.filter(atleta=atleta, año=año, mes=mes_num).first()
            if mensualidad:
                monto = float(mensualidad.monto_pagado)
                if mensualidad.exonerado:
                    estado_mes = "exonerado"
                elif monto >= 8:
                    estado_mes = "pagado"
                elif monto > 0:
                    estado_mes = "parcial"
                else:
                    estado_mes = "no_pagado"
            else:
                estado_mes = "no_pagado"

            if estado and estado_mes == estado:
                tiene_estado = True

        if estado and not tiene_estado and not atletas_filtrados_por_cedula:
            continue

        fila['Pagado'] = total_pagado
        fila['Pendiente'] = total_pendiente
        data.append(fila)

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Pagos')

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    ws.insert_rows(1, amount=4)

    # Logo más grande
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-olympo.jpeg')
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 100
        img.height = 100
        ws.add_image(img, 'A1')

    # Título y subtítulo
    ws['C1'] = "Escuela Deportiva OLYMPO"
    ws['C2'] = f"Reporte de Pagos - Año {año}"
    ws['C1'].font = Font(size=14, bold=True)
    ws['C2'].font = Font(size=12, bold=True)
    ws['C1'].alignment = Alignment(vertical="center")
    ws['C2'].alignment = Alignment(vertical="center")
    ws.merge_cells('C1:H1')
    ws.merge_cells('C2:H2')
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24

    # Encabezado de tabla
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    # Ancho personalizado: Categoría más estrecho
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        header = ws[f'{col_letter}5'].value
        if header == 'Categoría':
            ws.column_dimensions[col_letter].width = 13
        else:
            ws.column_dimensions[col_letter].width = 13

    # Estilo para celdas especiales
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        for cell in row:
            if cell.value == "E":
                cell.font = Font(bold=True, color="0000FF")
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="center")

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    response = HttpResponse(
        final_output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_pagos.xlsx"'
    return response

@solo_login
def exportar_pagos_pdf(request):
    from datetime import datetime
    año = int(request.GET.get('año', date.today().year))
    categoria = request.GET.get('categoria')
    nombre = request.GET.get('nombre')
    cedula = request.GET.get('cedula')
    estado = request.GET.get('estado')  # 'pagado', 'pendiente', 'parcial', 'exonerado'
    mes_inicio = int(request.GET.get('mes_inicio', 1))
    mes_fin = int(request.GET.get('mes_fin', 12))

    atletas = Atleta.objects.all()
    if categoria:
        atletas = atletas.filter(categoria=categoria)
    if nombre:
        atletas = atletas.filter(nombre__icontains=nombre)
    if cedula:
        atletas = atletas.filter(cedula__icontains=cedula)

    atletas_filtrados_por_cedula = False
    if cedula and atletas.count() == 1:
        atletas_filtrados_por_cedula = True

    meses = [(i, date(1900, i, 1).strftime('%b')) for i in range(mes_inicio, mes_fin + 1)]
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(letter))
    width, height = landscape(letter)

    # Para numerar páginas
    page_number = 1

    headers = ["Nombre", "Cédula", "Categoría"] + [m[1][:3] for m in meses] + ["Pagado", "Pendiente"]
    # Ajuste dinámico: si hay 12 meses, reduce un poco las columnas
    if len(meses) == 12:
        col_width = max(42, min(55, (width - 80) // len(headers)))
    else:
        col_width = max(45, min(60, (width - 100) // len(headers)))


    def imprimir_encabezado(pagina_inicial=False):
        nonlocal y
        y = height - 40
        x = 40
        if pagina_inicial:
            # Logo solo en la primera página
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-olympo.jpeg')
            if os.path.exists(logo_path):
                logo = ImageReader(logo_path)
                p.drawImage(logo, x, y - 40, width=80, height=80)

            # Título
            p.setFont("Helvetica-Bold", 16)
            p.drawString(x + 90, y, "Escuela Deportiva OLYMPO")
            p.setFont("Helvetica", 12)
            p.drawString(x + 90, y - 20, f"Reporte de Pagos - Año {año}")
            y -= 60
        else:
            y = height - 40
            p.setFont("Helvetica", 7)

        # Encabezado de tabla
        p.setFont("Helvetica-Bold", 8)
        for i, h in enumerate(headers):
            p.drawString(x + i * col_width, y, h)
        y -= 18
        p.setFont("Helvetica", 7)

    def imprimir_pie_pagina():
        fecha_export = datetime.now().strftime("%d/%m/%Y %H:%M")
        p.setFont("Helvetica", 8)
        p.drawRightString(width - 60, 25, f"Página {page_number}")
        p.drawString(40, 25, f"Exportado: {fecha_export}")
        p.setFont("Helvetica", 7)

    # Primera página con logo
    imprimir_encabezado(pagina_inicial=True)

    y = height - 40 - 60 - 18  # Ajusta y para la primera fila de datos

    for atleta in atletas:
        fila = [atleta.nombre, atleta.cedula, atleta.categoria]
        total_pagado = 0
        total_pendiente = 0
        pagos_mes = []
        tiene_estado = False

        for mes_num, _ in meses:
            mensualidad = Mensualidad.objects.filter(atleta=atleta, año=año, mes=mes_num).first()
            if mensualidad:
                monto = float(mensualidad.monto_pagado)
                if mensualidad.exonerado:
                    estado_mes = "exonerado"
                    pagos_mes.append("E")
                elif monto >= 8:
                    estado_mes = "pagado"
                    pagos_mes.append(f"{monto:.2f}")
                elif monto > 0:
                    estado_mes = "parcial"
                    pagos_mes.append(f"{monto:.2f}")
                else:
                    estado_mes = "no_pagado"
                    pagos_mes.append("—")
                if estado_mes == estado:
                    tiene_estado = True
                if estado_mes != "exonerado":
                    total_pagado += monto
                    if monto < 8:
                        total_pendiente += (8 - monto)
            else:
                pagos_mes.append("—")
                total_pendiente += 8
                if estado == "no_pagado":
                    tiene_estado = True

            # Si no hay ninguna mensualidad creada y estamos filtrando por pendiente,
            # igual debemos considerar que está pendiente.
            if not Mensualidad.objects.filter(atleta=atleta, año=año, mes__gte=mes_inicio, mes__lte=mes_fin).exists():
                if estado == "no_pagado":
                    tiene_estado = True

        if estado and not tiene_estado and not atletas_filtrados_por_cedula:
            continue

        fila.extend(pagos_mes)
        fila.append(f"{total_pagado:.2f}")
        fila.append(f"{total_pendiente:.2f}")

        for i, val in enumerate(fila):
            p.drawString(40 + i * col_width, y, str(val))
        y -= 14

        if y < 50:
            imprimir_pie_pagina()
            p.showPage()
            page_number += 1
            imprimir_encabezado(pagina_inicial=False)
            y -= 18  # Ajuste para la siguiente fila de datos

    # Pie de página en la última página
    imprimir_pie_pagina()

    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='reporte_pagos.pdf')

@solo_login
def reporte_atletas_view(request):
    categorias = ['U9', 'U11', 'U13', 'U15', 'U17', 'U19', 'U21', 'U23', 'Libre']
    atletas = Atleta.objects.all()
    
    cedula = request.GET.get('cedula', '').strip()
    nombre = request.GET.get('nombre', '').strip()
    categoria = request.GET.get('categoria', '')
    edad_min = request.GET.get('edad_min')
    edad_max = request.GET.get('edad_max')
    sexo = request.GET.get('sexo', '').strip()

    if cedula:
        atletas = atletas.filter(cedula__icontains=cedula)
    if nombre:
        atletas = atletas.filter(nombre__icontains=nombre) | atletas.filter(apellido__icontains=nombre)
    if categoria:
        atletas = atletas.filter(categoria=categoria)
    if edad_min:
        atletas = atletas.filter(fecha_nacimiento__year__lte=date.today().year - int(edad_min))
    if edad_max:
        atletas = atletas.filter(fecha_nacimiento__year__gte=date.today().year - int(edad_max))
    if sexo:
        atletas = atletas.filter(sexo__iexact=sexo)

    años_disponibles = sorted({a.fecha_registro.year for a in Atleta.objects.all()})

    # PAGINACIÓN

    paginator = Paginator(atletas, 15)  # 15 atletas por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        'atletas': page_obj,
        'años_disponibles': años_disponibles,
        'categorias': categorias,
        'filtros': {
            'cedula': cedula,
            'nombre': nombre,
            'categoria': categoria,
            'edad_min': edad_min,
            'edad_max': edad_max,
            'sexo': sexo,
        },
        'page_obj': page_obj,
    }
    return render(request, 'reportes/reporte_atletas.html', context)

def aplicar_filtros(request):
    atletas = Atleta.objects.all()

    cedula = request.GET.get('cedula', '').strip()
    nombre = request.GET.get('nombre', '').strip()
    categoria = request.GET.get('categoria', '')
    edad_min = request.GET.get('edad_min')
    edad_max = request.GET.get('edad_max')
    sexo = request.GET.get('sexo', '')

    if cedula:
        atletas = atletas.filter(cedula__icontains=cedula)
    if nombre:
        atletas = atletas.filter(nombre__icontains=nombre) | atletas.filter(apellido__icontains=nombre)
    if categoria:
        atletas = atletas.filter(categoria=categoria)
    if edad_min:
        atletas = atletas.filter(fecha_nacimiento__year__lte=date.today().year - int(edad_min))
    if edad_max:
        atletas = atletas.filter(fecha_nacimiento__year__gte=date.today().year - int(edad_max))
    if sexo:
        atletas = atletas.filter(sexo__iexact=sexo)

    return atletas

@solo_login
def reporte_atletas_excel(request):

    atletas = aplicar_filtros(request)
    
    data = []
    for a in atletas:
        data.append({
            'Nombre y Apellido': f"{a.nombre} {a.apellido}",
            'Cédula': a.cedula,
            'Categoría': a.categoria,
            'Edad': a.calcular_edad(),
            'Sexo': a.sexo,
            'Teléfono': a.telefono,
            'Dirección': a.direccion,
        })

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Atletas')

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    ws.insert_rows(1, amount=4)

    # Logo y encabezados
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-olympo.jpeg')
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 100
        img.height = 100
        ws.add_image(img, 'A1')

    ws['C1'] = "Escuela Deportiva OLYMPO"
    ws['C2'] = "Reporte de Atletas"
    ws['C1'].font = Font(size=14, bold=True)
    ws['C2'].font = Font(size=12, bold=True)
    ws['C1'].alignment = Alignment(vertical="center")
    ws['C2'].alignment = Alignment(vertical="center")
    ws.merge_cells('C1:H1')
    ws.merge_cells('C2:H2')
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24

    # Estilo encabezado tabla
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    # Anchos personalizados y alineación por columna
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        header = ws[f'{col_letter}5'].value

        if header == 'Nombre y Apellido':
            ws.column_dimensions[col_letter].width = 30
            align = Alignment(horizontal="left")
        elif header == 'Cédula':
            ws.column_dimensions[col_letter].width = 16
            align = Alignment(horizontal="center")
        elif header == 'Categoría':
            ws.column_dimensions[col_letter].width = 14
            align = Alignment(horizontal="center")
        elif header == 'Edad':
            ws.column_dimensions[col_letter].width = 10
            align = Alignment(horizontal="center")
        elif header == 'Sexo':
            ws.column_dimensions[col_letter].width = 10
            align = Alignment(horizontal="center")
        elif header == 'Teléfono':
            ws.column_dimensions[col_letter].width = 18
            align = Alignment(horizontal="center")
        elif header == 'Dirección':
            ws.column_dimensions[col_letter].width = 35
            align = Alignment(horizontal="left")
        else:
            align = Alignment(horizontal="center")

        # Aplicar alineación a toda la columna
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = align

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    response = HttpResponse(
        final_output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_atletas.xlsx"'
    return response

@solo_login
def reporte_atletas_pdf(request):
    atletas = aplicar_filtros(request)
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    col_widths = [100, 60, 60, 60, 60, 80, 140]
    headers = ["Nombre y Apellido", "Cédula", "Categoría", "Edad", "Sexo", "Teléfono", "Dirección"]

    y = height - 40
    page_number = 1

    def encabezado(pagina_inicial=False):
        nonlocal y
        y = height - 40
        x = 40

        if pagina_inicial:
            # Logo
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-olympo.jpeg')
            if os.path.exists(logo_path):
                logo = ImageReader(logo_path)
                p.drawImage(logo, x, y - 40, width=80, height=80)

            # Título
            p.setFont("Helvetica-Bold", 16)
            p.drawString(x + 90, y, "Escuela Deportiva OLYMPO")
            p.setFont("Helvetica", 12)
            p.drawString(x + 90, y - 20, "Reporte de Atletas")
            y -= 60
        else:
            y = height - 40

        # Encabezado de tabla
        p.setFont("Helvetica-Bold", 8)
        curr_x = x
        for i, header in enumerate(headers):
            p.drawString(curr_x, y, header)
            curr_x += col_widths[i]
        y -= 18
        p.setFont("Helvetica", 7)

    def pie_pagina():
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        p.setFont("Helvetica", 8)
        p.drawRightString(width - 40, 25, f"Página {page_number}")
        p.drawString(40, 25, f"Exportado: {fecha}")

    # Primer encabezado
    encabezado(pagina_inicial=True)

    # Datos
    for atleta in atletas:
        fila = [
            f"{atleta.nombre} {atleta.apellido}",
            atleta.cedula,
            atleta.categoria,
            atleta.calcular_edad(),
            atleta.sexo,
            atleta.telefono,
            atleta.direccion,

        ]

        curr_x = 40
        for i, dato in enumerate(fila):
            p.drawString(curr_x, y, str(dato))
            curr_x += col_widths[i]
        y -= 14

        if y < 50:
            pie_pagina()
            p.showPage()
            page_number += 1
            encabezado(pagina_inicial=False)
            y -= 18

    pie_pagina()
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='reporte_atletas.pdf')

@solo_login
def reporte_equipos(request):
    nombre = request.GET.get("nombre", "").strip()
    sexo_param = request.GET.get("sexo", "").strip()

    equipos = Equipo.objects.all().annotate(num_atletas=Count("atletas"))

    # filtro por nombre
    if nombre:
        equipos = equipos.filter(nombre__icontains=nombre)

    # filtro por sexo
    sexo_normalizado = ""
    if sexo_param:
        s = sexo_param.lower()
        if s in ("masculino", "femenino", "mixto"):
            sexo_normalizado = s
        elif s == "m":
            sexo_normalizado = "masculino"
        elif s == "f":
            sexo_normalizado = "femenino"
    if sexo_normalizado:
        equipos = equipos.filter(sexo_equipo=sexo_normalizado)

    paginator = Paginator(equipos, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "page_obj": page_obj,
        "nombre": nombre,
        "sexo": sexo_normalizado or sexo_param,

    }
    return render(request, "reportes/reporte_equipos.html", context)

@solo_login
def exportar_equipo_pdf(request, equipo_id):
    equipo = get_object_or_404(Equipo, id=equipo_id)
    atletas = equipo.atletas.all()

    # response
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="equipo_{equipo.nombre}.pdf"'

    # estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TituloPrincipal", fontSize=20, leading=24, alignment=1,
                              textColor=colors.HexColor("#2C3E50"), spaceAfter=16))
    styles.add(ParagraphStyle(name="SubTitulo", fontSize=14, leading=18,
                              textColor=colors.HexColor("#2980B9"), spaceAfter=10))
    styles.add(ParagraphStyle(name="NormalCustom", fontSize=11, leading=13, spaceAfter=4))

    # ruta logo
    logo_path = os.path.join(settings.BASE_DIR, "static", "img", "logo-olympo.jpeg")

    # encabezado / pie página
    def header_footer(canvas, doc):
        width, height = letter
        canvas.setStrokeColor(colors.HexColor("#2C3E50"))
        canvas.setLineWidth(1)
        canvas.line(40, height - 70, width - 40, height - 70)

        # logo
        if os.path.exists(logo_path):
            canvas.drawImage(logo_path, 40, height - 65, width=70, height=70, preserveAspectRatio=True)

        # título encabezado
        canvas.setFont("Helvetica-Bold", 14)
        canvas.setFillColor(colors.HexColor("#2C3E50"))
        canvas.drawString(120, height - 40, "Escuela Deportiva OLYMPO")

        # pie página
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.black)
        canvas.drawString(40, 30, f"Exportado: {fecha}")
        canvas.drawRightString(width - 40, 30, f"Página {doc.page}")

    # documento
    doc = SimpleDocTemplate(response, pagesize=letter,
                            rightMargin=40, leftMargin=40,
                            topMargin=90, bottomMargin=40)

    elements = []

    # título principal
    elements.append(Paragraph("Reporte del Equipo", styles["TituloPrincipal"]))
    elements.append(Spacer(1, 12))

    # encabezado de la sección de información del equipo
    elements.append(Paragraph("Información del Equipo", styles["SubTitulo"]))
    elements.append(Spacer(1, 4))

    # info del equipo (párrafos pegados a la izquierda)
    elements.append(Paragraph(f"<b>Nombre:</b> {equipo.nombre}", styles["NormalCustom"]))
    elements.append(Paragraph(f"<b>Categoría:</b> {equipo.categoria}", styles["NormalCustom"]))
    elements.append(Paragraph(f"<b>Sexo:</b> {equipo.get_sexo_equipo_display()}", styles["NormalCustom"]))
    elements.append(Paragraph(f"<b>Cantidad de Atletas:</b> {equipo.atletas.count()}", styles["NormalCustom"]))
    elements.append(Spacer(1, 20))

    # tabla atletas
    if atletas.exists():
        elements.append(Paragraph("Listado de Atletas", styles["SubTitulo"]))
        data = [["Nombre", "Cédula", "Edad", "Teléfono", "Sexo"]]
        for a in atletas:
            data.append([
                f"{a.nombre} {a.apellido}",
                a.cedula,
                f"{a.calcular_edad()} años",
                a.telefono,
                a.sexo.title()
            ])
        # 👇 aumentamos la última columna (Sexo) de 50 → 70
        table = Table(data, hAlign="CENTER", colWidths=[140, 90, 60, 100, 70])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("Este equipo no tiene atletas asignados.", styles["NormalCustom"]))

    # construir pdf
    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)
    return response

# LISTAR
@solo_login
def lista_administradores(request):
    administradores = Administrador.objects.select_related("usuario").all().order_by("apellido", "nombre")
    paginator = Paginator(administradores, 8)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "administradores/lista.html", {
        "page_obj": page_obj,
        "administradores": page_obj.object_list,
    })


# AGREGAR
@solo_login
def agregar_administrador(request):
    if request.method == "POST":
        form = AdministradorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("lista_administradores")
    else:
        form = AdministradorForm()

    return render(request, "administradores/agregar.html", {
        "form": form
    })


# DETALLE
@solo_login
def detalle_administrador(request, pk):
    administrador = get_object_or_404(Administrador.objects.select_related("usuario"), pk=pk)
    return render(request, "administradores/detalle.html", {
        "administrador": administrador
    })


# EDITAR
@solo_login
def editar_administrador(request, administrador_id):
    administrador = get_object_or_404(Administrador, id=administrador_id)

    if request.method == "POST":
        form = AdministradorForm(request.POST, instance=administrador)
        if form.is_valid():
            form.save()
            return redirect("detalle_administrador", pk=administrador.id)
    else:
        form = AdministradorForm(instance=administrador)

    return render(request, "administradores/editar.html", {
        "form": form,
        "administrador": administrador,
    })


# ELIMINAR
@solo_login
def eliminar_administrador(request, administrador_id):
    administrador = get_object_or_404(Administrador, id=administrador_id)

    if request.method == "POST":
        # Importante: borrar también el User asociado para no dejar basura
        user = administrador.usuario
        administrador.delete()
        if user:
            user.delete()

        return redirect("lista_administradores")

    return render(request, "administradores/eliminar.html", {
        "administrador": administrador
    })